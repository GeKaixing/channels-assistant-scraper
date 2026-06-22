"""
视频号助手 — 三张表一键导出
数据中心: 关注者数据 + 视频数据 + 单篇视频 → Excel（无中间文件）

跨平台兼容: macOS / Linux / Windows

用法: python3 templates/scrape_all_three.py
输出: ~/Desktop/视频号助手数据/关注者数据.xlsx
       ~/Desktop/视频号助手数据/视频日趋势.xlsx
       ~/Desktop/视频号助手数据/单篇视频.xlsx
"""

import os, sys, asyncio, openpyxl, tempfile
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.async_api import async_playwright

SKILL_DIR = os.path.dirname(os.path.abspath(__file__))
STATE_PATH = os.path.join(SKILL_DIR, "..", "assets", "channels_sessions.json")

# ── 跨平台路径 ──
def get_desktop_path(sub_dir=None):
    home = os.path.expanduser("~")
    if sys.platform == "darwin":
        base = os.path.join(home, "Desktop")
    elif sys.platform == "win32":
        base = os.path.join(home, "Desktop")
    else:
        base = os.environ.get("XDG_DESKTOP_DIR", os.path.join(home, "Desktop"))
    if not os.path.exists(base):
        base = home
    return os.path.join(base, sub_dir) if sub_dir else base

OUT_DIR = get_desktop_path("视频号助手数据")

fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr = Font(bold=True, color="FFFFFF", size=11)


def style_headers(ws, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")


def to_int(s):
    try:
        return int(s)
    except (ValueError, TypeError):
        return s


async def get_coords(page, text):
    """Find sidebar element click coordinates by exact text match"""
    return await page.evaluate(f"""() => {{
        const all = document.querySelectorAll('a, li, span, div');
        for (const el of all) {{
            if (el.textContent.trim() === '{text}' && el.offsetParent !== null) {{
                const r = el.getBoundingClientRect();
                return {{x: Math.round(r.x + r.width / 2), y: Math.round(r.y + r.height / 2)}};
            }}
        }}
        return null;
    }}""")


async def click_download_btn(frame):
    """Click '下载表格' inside a micro-app iframe"""
    return await frame.evaluate("""() => {
        const all = document.querySelectorAll('button, a, span, div');
        for (const el of all) {
            if (el.textContent.trim() === '下载表格' && el.offsetParent !== null) {
                el.click();
                return true;
            }
        }
        return false;
    }""")


def extract_tsv(text):
    """Extract tab-separated data rows (lines starting with year) from iframe text"""
    rows = []
    for l in text.split("\n"):
        l = l.strip()
        if l and (l.startswith("2026/") or l.startswith("2025/")):
            rows.append(l.split("\t"))
    return rows


async def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            storage_state=STATE_PATH, viewport={"width": 1920, "height": 1080}
        )
        page = await ctx.new_page()

        await page.goto(
            "https://channels.weixin.qq.com/platform",
            wait_until="domcontentloaded",
            timeout=30_000,
        )
        await asyncio.sleep(3)

        if "/login" in page.url:
            print("SESSION EXPIRED — 请重新运行 login_helper.py")
            await browser.close()
            return

        # ── 1. 关注者数据 ──
        dc = await get_coords(page, "数据中心")
        if dc:
            await page.mouse.click(dc["x"], dc["y"])
            await asyncio.sleep(2)

        fans = await get_coords(page, "关注者数据")
        if fans:
            await page.mouse.click(fans["x"], fans["y"])
            await asyncio.sleep(6)

        fans_text = ""
        for f in page.frames:
            if "micro/statistic/follower" in f.url:
                await click_download_btn(f)
                await asyncio.sleep(3)
                fans_text = await f.evaluate('document.body.innerText or ""')
                break

        # ── 2. 视频数据 ──
        video = await get_coords(page, "视频数据")
        if video:
            await page.mouse.click(video["x"], video["y"])
            await asyncio.sleep(6)

        video_text = ""
        single_text = ""
        for f in page.frames:
            if "micro/statistic/post" in f.url:
                await click_download_btn(f)
                await asyncio.sleep(3)
                video_text = await f.evaluate('document.body.innerText or ""')

                # Switch to 单篇视频 tab
                await f.evaluate("""() => {
                    const all = document.querySelectorAll('span, a, div, li');
                    for (const el of all) {
                        if (el.textContent.trim() === '单篇视频' && el.offsetParent !== null) {
                            el.click();
                            return true;
                        }
                    }
                    return false;
                }""")
                await asyncio.sleep(5)

                await click_download_btn(f)
                await asyncio.sleep(3)
                single_text = await f.evaluate('document.body.innerText or ""')
                break

        await browser.close()

    # ===== 生成 Excel =====

    # ── 关注者数据.xlsx ──
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "关注者数据"
    style_headers(ws1, ["日期", "净增关注", "新增关注", "取消关注", "关注者总数"])
    ri = 2
    for cols in extract_tsv(fans_text):
        if len(cols) >= 5:
            for ci in range(5):
                ws1.cell(row=ri, column=ci + 1, value=to_int(cols[ci]))
            ri += 1
    for c in "ABCDE":
        ws1.column_dimensions[c].width = 14
    ws1.freeze_panes = "A2"
    fp1 = os.path.join(OUT_DIR, "关注者数据.xlsx")
    wb1.save(fp1)
    print(f"✅ {fp1} ({ri - 2} 天)")

    # ── 视频日趋势.xlsx ──
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "视频日趋势"
    style_headers(ws2, ["日期", "播放", "点赞", "评论", "分享", "关注"])
    ri = 2
    for cols in extract_tsv(video_text):
        clean = [c for c in cols if c.strip()]
        if len(clean) >= 6:
            for ci in range(6):
                ws2.cell(row=ri, column=ci + 1, value=to_int(clean[ci]))
            ri += 1
    for c in "ABCDEF":
        ws2.column_dimensions[c].width = 12
    ws2.freeze_panes = "A2"
    fp2 = os.path.join(OUT_DIR, "视频日趋势.xlsx")
    wb2.save(fp2)
    print(f"✅ {fp2} ({ri - 2} 天)")

    # ── 单篇视频.xlsx ──
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "单篇视频"
    style_headers(
        ws3,
        [
            "视频标题", "发布时间", "完播率", "平均播放时长",
            "播放量", "评论", "关注", "分享",
            "转发聊天和朋友圈", "设为铃声", "设为状态", "设为朋友圈封面",
        ],
    )
    lines = single_text.split("\n")
    ri = 2
    for i, l in enumerate(lines):
        l = l.strip()
        if l.startswith("2026/") or l.startswith("2025/"):
            title = lines[i - 1].strip() if i > 0 else ""
            cols = l.split("\t")
            ws3.cell(row=ri, column=1, value=title[:60])
            for ci, val in enumerate(cols[:11]):
                ws3.cell(row=ri, column=ci + 2, value=to_int(val))
            ri += 1

    ws3.column_dimensions["A"].width = 50
    for c in "BCDEFGHIJKL":
        ws3.column_dimensions[c].width = 14
    ws3.freeze_panes = "A2"
    fp3 = os.path.join(OUT_DIR, "单篇视频.xlsx")
    wb3.save(fp3)
    print(f"✅ {fp3} ({ri - 2} 条)")


if __name__ == "__main__":
    asyncio.run(main())
