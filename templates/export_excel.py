"""
视频号助手 - JSON 转 Excel
读取已保存的 all_api_data.json + dashboard_dom.json，生成多 Sheet Excel 报告

用法: python3 templates/export_excel.py
输出: ~/Desktop/视频号助手数据/视频号助手数据.xlsx

跨平台兼容: macOS / Linux / Windows
"""

import os, sys, json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SKILL_DIR = os.path.dirname(os.path.abspath(__file__))

def get_desktop_path(sub_dir=None):
    home = os.path.expanduser("~")
    if sys.platform == "darwin":
        base = os.path.join(home, "Desktop")
    elif sys.platform == "win32":
        base = os.path.join(home, "Desktop")
    else:
        base = os.environ.get("XDG_DESKTOP_DIR", os.path.join(home, "Desktop"))
    if not os.path.exists(base):
        base = home
    return os.path.join(base, sub_dir) if sub_dir else base

DATA_DIR = get_desktop_path("视频号助手数据")

fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(bold=True, color="FFFFFF", size=11)


def write_header(ws, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main():
    os.makedirs(DATA_DIR, exist_ok=True)

    with open(os.path.join(DATA_DIR, "all_api_data.json")) as f:
        api = json.load(f)
    with open(os.path.join(DATA_DIR, "dashboard_dom.json")) as f:
        dom = json.load(f)

    wb = openpyxl.Workbook()
    lines = dom.get("lines", [])
    parsed = dom.get("parsed", {})

    # === Sheet 1: 概览 ===
    ws1 = wb.active
    ws1.title = "概览"
    write_header(ws1, ["字段", "值"])
    overview = [
        ("视频号ID", parsed.get("finder_id", "")),
        ("关注者", parsed.get("followers", "")),
        ("视频数", parsed.get("video_count", "")),
        ("全页文本行数", len(lines)),
        ("页面URL", parsed.get("url", "")),
    ]
    for ri, (k, v) in enumerate(overview, 2):
        ws1.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=ri, column=2, value=v)
    yesterday = parsed.get("yesterday", {})
    if yesterday:
        row = len(overview) + 3
        ws1.cell(row=row, column=1, value="昨日数据").font = Font(bold=True, size=12)
        for k, v in yesterday.items():
            row += 1
            ws1.cell(row=row, column=1, value=k)
            ws1.cell(row=row, column=2, value=v)
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 40

    # === Sheet 2: 视频列表 ===
    ws2 = wb.create_sheet("视频列表")
    write_header(ws2, ["发布时间", "播放量", "点赞", "评论", "转发", "收藏", "时长(秒)", "视频ID"])
    pl = api.get("post/post_list", {})
    posts = pl.get("data", {}).get("list", [])
    for ri, p in enumerate(posts, 2):
        ts = datetime.fromtimestamp(p.get("createTime", 0)).strftime("%Y-%m-%d %H:%M")
        media = (p.get("desc", {}) or {}).get("media", [{}])[0] if p.get("desc") else {}
        ws2.cell(row=ri, column=1, value=ts)
        ws2.cell(row=ri, column=2, value=p.get("readCount", 0))
        ws2.cell(row=ri, column=3, value=p.get("likeCount", 0))
        ws2.cell(row=ri, column=4, value=p.get("commentCount", 0))
        ws2.cell(row=ri, column=5, value=p.get("forwardCount", 0))
        ws2.cell(row=ri, column=6, value=p.get("favCount", 0))
        ws2.cell(row=ri, column=7, value=media.get("videoPlayLen", ""))
        ws2.cell(row=ri, column=8, value=p.get("objectId", "")[-20:])
    for col, w in {"A": 18, "B": 10, "C": 8, "D": 8, "E": 8, "F": 8, "G": 10, "H": 30}.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # === Sheet 3: 粉丝趋势 ===
    ws3 = wb.create_sheet("粉丝趋势")
    write_header(ws3, ["指标", "前期值", "后期值"])
    ft = api.get("statistic/fans_trend", {}).get("data", {}).get("data", {})
    trend_items = [
        ("新增", ft.get("add", [])),
        ("减少", ft.get("reduce", [])),
        ("净增", ft.get("netAdd", [])),
        ("总数", ft.get("total", [])),
    ]
    for ri, (name, vals) in enumerate(trend_items, 2):
        ws3.cell(row=ri, column=1, value=name).font = Font(bold=True)
        ws3.cell(row=ri, column=2, value=vals[0] if len(vals) > 0 else "")
        ws3.cell(row=ri, column=3, value=vals[1] if len(vals) > 1 else "")
    tabtypes = ft.get("fansDataByTabtype", [])
    if tabtypes:
        row = len(trend_items) + 3
        ws3.cell(row=row, column=1, value="按来源分类").font = Font(bold=True, size=12)
        row += 1
        write_header(ws3, ["来源", "新增(前)", "新增(后)", "净增(前)", "净增(后)", "总数(前)", "总数(后)"])
        for t in tabtypes:
            row += 1
            ws3.cell(row=row, column=1, value=t.get("tabTypeName", ""))
            for ci, key in enumerate(["add", "netAdd", "total"], 2):
                v = t.get(key, [])
                ws3.cell(row=row, column=ci * 2 - 2, value=v[0] if len(v) > 0 else "")
                ws3.cell(row=row, column=ci * 2 - 1, value=v[1] if len(v) > 1 else "")
    ws3.column_dimensions["A"].width = 14
    for c in "BCDEFGH":
        ws3.column_dimensions[c].width = 12

    # === Sheet 4: 作品数据 ===
    ws4 = wb.create_sheet("作品数据")
    write_header(ws4, ["指标", "前期值", "后期值"])
    pt = api.get("statistic/new_post_total_data", {}).get("data", {}).get("data", {})
    tabs = pt.get("dataByTabtype", [])
    LABEL_MAP = {"browse": "播放", "like": "点赞", "comment": "评论", "forward": "转发", "fav": "收藏"}
    for t in tabs:
        if t.get("tabType") == 3:
            data = t.get("data", {})
            for ri, key in enumerate(["browse", "like", "comment", "forward", "fav"], 2):
                v = data.get(key, [])
                ws4.cell(row=ri, column=1, value=LABEL_MAP.get(key, key))
                ws4.cell(row=ri, column=2, value=int(v[0]) if len(v) > 0 and v[0] else "")
                ws4.cell(row=ri, column=3, value=int(v[1]) if len(v) > 1 and v[1] else "")
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 14

    # === Sheet 5: 小店信息 ===
    ws5 = wb.create_sheet("小店信息")
    write_header(ws5, ["字段", "值"])
    ec = api.get("shop/get_finder_ec_info_for_opening_page", {}).get("data", {}).get("data", {})
    shop_info = ec.get("shopInfo", {}) or {}
    shop_rows = [
        ("商户状态", {1: "未开通", 2: "已开通"}.get(ec.get("merchant"), ec.get("merchant"))),
        ("带货者状态", {1: "已开通"}.get(ec.get("promoter"), ec.get("promoter"))),
        ("小店名称", shop_info.get("shopName", "")),
        ("小店AppID", shop_info.get("appid", "")),
        ("视频号小店名", ec.get("windowName", "")),
    ]
    for ri, (k, v) in enumerate(shop_rows, 2):
        ws5.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws5.cell(row=ri, column=2, value=v)
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 50

    # === Sheet 6: 通知 ===
    ws6 = wb.create_sheet("通知")
    write_header(ws6, ["时间", "标题", "内容", "已读"])
    notifs = api.get("notification/notification_list", {}).get("data", {}).get("data", {}).get("list", [])
    for ri, n in enumerate(notifs, 2):
        ts = datetime.fromtimestamp(n.get("createTime", 0)).strftime("%Y-%m-%d %H:%M")
        ws6.cell(row=ri, column=1, value=ts)
        ws6.cell(row=ri, column=2, value=n.get("title", ""))
        ws6.cell(row=ri, column=3, value=n.get("content", ""))
        ws6.cell(row=ri, column=4, value="是" if n.get("isRead") else "否")
    for col, w in {"A": 18, "B": 30, "C": 60, "D": 8}.items():
        ws6.column_dimensions[col].width = w

    # === Sheet 7: API端点 ===
    ws7 = wb.create_sheet("API端点")
    write_header(ws7, ["端点", "状态码", "数据类型"])
    for ri, (ep, info) in enumerate(sorted(api.items()), 2):
        ws7.cell(row=ri, column=1, value=ep)
        ws7.cell(row=ri, column=2, value=info.get("status", ""))
        ws7.cell(row=ri, column=3, value=type(info.get("data", {})).__name__)
    ws7.column_dimensions["A"].width = 50
    ws7.column_dimensions["B"].width = 10
    ws7.column_dimensions["C"].width = 15

    fp = os.path.join(DATA_DIR, "视频号助手数据.xlsx")
    wb.save(fp)
    print(f"✅ 视频号助手数据.xlsx")
    print(f"   {fp}")
    print(f"   共 {len(wb.sheetnames)} 个 Sheet: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
